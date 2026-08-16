from __future__ import annotations

from datetime import datetime
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import OrderResult, ProcessingRecord


def report_path_for_run(run_dir: Path) -> Path:
    return run_dir / f"Report_{run_dir.name}.xlsx"


class ReportWriter:
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def __init__(self, run_dir: Path, logger):
        self.run_dir = run_dir
        self.logger = logger

    def write(
        self,
        discovered: list[OrderResult],
        records: list[ProcessingRecord],
        started_at: datetime,
        finished_at: datetime,
        settings_snapshot: dict[str, Any],
    ) -> Path:
        report_path = report_path_for_run(self.run_dir)
        counts: dict[str, int] = {}
        for record in records:
            counts[record.status] = counts.get(record.status, 0) + 1
        workbook = Workbook()
        workbook.remove(workbook.active)
        summary = workbook.create_sheet("Summary")
        summary.append(["Metric", "Value"])
        summary_rows = [
            ("Started", started_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Finished", finished_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Duration seconds", round((finished_at - started_at).total_seconds(), 2)),
            ("Discovered PDF orders", len(discovered)),
            ("Collected", counts.get("collected", 0)),
            ("IRT duplicates skipped", counts.get("duplicate", 0)),
            ("IRT-backed consolidated copies skipped", counts.get("consolidated_duplicate", 0)),
            ("Local duplicates skipped", counts.get("local_duplicate", 0)),
            ("Content duplicates removed", counts.get("content_duplicate", 0)),
            ("Errors", counts.get("error", 0)),
            ("Cancelled", counts.get("cancelled", 0)),
            ("IRT court code", settings_snapshot.get("irt_court_code", "")),
        ]
        for row in summary_rows:
            summary.append(row)
        self._style(summary)

        collected = [row.as_dict() for row in records if row.status == "collected"]
        duplicates = [
            row.as_dict()
            for row in records
            if row.status
            in {"duplicate", "consolidated_duplicate", "local_duplicate", "content_duplicate"}
        ]
        errors = [row.as_dict() for row in records if row.status in {"error", "cancelled"}]
        self._add_filenames_sheet(workbook, records)
        self._add_records_sheet(workbook, "Collected", collected)
        self._add_records_sheet(workbook, "Duplicates", duplicates)
        self._add_records_sheet(workbook, "Errors", errors)
        self._add_records_sheet(workbook, "Discovered", [row.as_dict() for row in discovered])
        workbook.save(report_path)
        self.logger.info("Report saved: %s", report_path.name)
        return report_path

    def _add_filenames_sheet(
        self, workbook, records: list[ProcessingRecord]
    ) -> None:
        sheet = workbook.create_sheet("Filenames")
        sheet.append(["Filename"])
        for record in records:
            if record.status == "collected" and record.target_filename:
                sheet.append([record.target_filename])
        self._style(sheet)

    def _add_records_sheet(self, workbook, name: str, rows: list[dict[str, Any]]) -> None:
        sheet = workbook.create_sheet(name)
        if not rows:
            sheet.append(["Status"])
            sheet.append([f"No {name.lower()} records"])
            self._style(sheet)
            return
        keys: list[str] = []
        for row in rows:
            for key in row:
                if key not in keys:
                    keys.append(key)
        sheet.append([self._label(key) for key in keys])
        for row in rows:
            sheet.append([self._cell_value(row.get(key, "")) for key in keys])
        self._style(sheet)

    @staticmethod
    def _cell_value(value: Any) -> Any:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False)
        return value

    @staticmethod
    def _label(key: str) -> str:
        return key.replace("_", " ").title().replace("Irt", "IRT").replace("Url", "URL")

    def _style(self, sheet) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        for column_cells in sheet.columns:
            max_length = max((len(str(cell.value or "")) for cell in column_cells), default=0)
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max_length + 2, 60)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
